from openpyxl.reader.excel import load_workbook
from collections import defaultdict


class Data:
    @staticmethod
    def getTestData(testcase):

        L=[]
        book =load_workbook("C:\\Users\\krist\\Desktop\\mydata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i,column=8).value == testcase:
                Dict = {}
                for j in range(1,8):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value or ""
                L.append(Dict)

        print(Dict)
        return L


#print(Data.getTestData("NOEMAIL"))

